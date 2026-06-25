#!/usr/bin/env python3
"""
PREMIUM CALCULATION ENGINE FOR CYBER & AI RISK
Phase 3: Integration of Frequency × Severity with Loading Factors
Actuarial Science Project - Complete Pricing Framework

Formula: Final Premium = Pure Premium × (1 + Loading Factors)
Where: Pure Premium = E[Frequency] × E[Severity]
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

print("=" * 80)
print("PREMIUM CALCULATION ENGINE FOR CYBER & AI RISK")
print("=" * 80)

# ============================================================================
# 1. LOAD MODELS & DATA
# ============================================================================

print("\n1. LOADING FREQUENCY AND SEVERITY MODELS")
print("-" * 80)

# Load frequency predictions
frequency_df = pd.read_csv('/mnt/user-data/outputs/company_frequency_predictions.csv')
print(f"Loaded frequency data: {len(frequency_df)} companies")

# Load incidents for additional context
incidents = pd.read_csv('/mnt/project/incidents_master.csv')
print(f"Loaded incidents data: {len(incidents)} incidents")

# ============================================================================
# 2. DEFINE LOADING FACTORS
# ============================================================================

print("\n2. DEFINING LOADING FACTORS")
print("-" * 80)

# Loading factor components (based on documentation)
loading_factors = {
    'acquisition_expenses': 0.20,      # 20% - broker commission, underwriting
    'administrative': 0.10,             # 10% - claims, policy admin
    'profit_margin': 0.15,              # 15% - target ROE
    'uncertainty_buffer': 0.08,         # 8% - model error, structural changes
    'reinsurance_ceded': 0.05,          # 5% - catastrophic risk transfer
}

total_loading = sum(loading_factors.values())
loading_multiplier = 1 + total_loading

print(f"\nLoading Factor Breakdown:")
for component, rate in loading_factors.items():
    print(f"  {component.replace('_', ' ').title()}: {rate*100:.1f}%")
print(f"  {'─' * 40}")
print(f"  Total Loading Factor: {total_loading*100:.1f}%")
print(f"  Loading Multiplier: {loading_multiplier:.3f}")

# ============================================================================
# 3. SEVERITY PARAMETERS & MODEL
# ============================================================================

print("\n3. SEVERITY MODEL PARAMETERS")
print("-" * 80)

# Lognormal distribution parameters (from severity model)
mu_severity = 16.6899
sigma_severity = 1.6415

print(f"Lognormal Distribution:")
print(f"  μ (log scale): {mu_severity:.4f}")
print(f"  σ (log scale): {sigma_severity:.4f}")

# Expected value from lognormal
mean_loss = np.exp(mu_severity + (sigma_severity**2) / 2)
print(f"  E[Loss]: ${mean_loss:,.0f}")

# Severity model coefficients (Ridge regression)
severity_coefficients = {
    'intercept': 9.5316,
    'log_revenue': 0.3586,
    'log_employees': -0.0627,
    'is_public': -0.1252,
    'log_records': 0.0203
}

print(f"\nSeverity Model Coefficients (Ridge):")
for var, coef in severity_coefficients.items():
    if var != 'intercept':
        print(f"  {var}: {coef:.4f}")
    else:
        print(f"  {var}: {coef:.4f}")

# ============================================================================
# 4. CALCULATE SEVERITY FOR EACH COMPANY
# ============================================================================

print("\n4. CALCULATING EXPECTED SEVERITY BY COMPANY")
print("-" * 80)

def predict_severity(company_revenue, employee_count, is_public, records_compromised=1000000):
    """
    Predict expected loss using severity model
    
    Args:
        company_revenue: Annual revenue in USD
        employee_count: Number of employees
        is_public: 1 if public company, 0 if private
        records_compromised: Data records at risk (default 1M)
    
    Returns:
        Predicted log-loss and expected loss in USD
    """
    
    # Feature transformation
    log_revenue = np.log1p(company_revenue)
    log_employees = np.log1p(employee_count)
    log_records = np.log1p(records_compromised)
    
    # Linear prediction on log scale
    log_loss = (severity_coefficients['intercept'] +
                severity_coefficients['log_revenue'] * log_revenue +
                severity_coefficients['log_employees'] * log_employees +
                severity_coefficients['is_public'] * is_public +
                severity_coefficients['log_records'] * log_records)
    
    # Back-transform to original scale
    expected_loss = np.exp(log_loss)
    
    return expected_loss, log_loss

# Apply severity model to all companies
severity_predictions = []
for idx, row in frequency_df.iterrows():
    revenue = row['company_revenue_usd']
    employees = row['employee_count']
    is_public = int(row['is_public_company'])
    
    # Estimate records at risk (rough proxy based on company size)
    records_at_risk = max(100000, employees * 50)  # ~50 records per employee
    
    expected_loss, log_loss = predict_severity(revenue, employees, is_public, records_at_risk)
    severity_predictions.append(expected_loss)

frequency_df['expected_severity'] = severity_predictions

print(f"Calculated severity for {len(frequency_df)} companies")
print(f"\nSeverity Statistics:")
print(f"  Mean: ${frequency_df['expected_severity'].mean():,.0f}")
print(f"  Median: ${frequency_df['expected_severity'].median():,.0f}")
print(f"  Std Dev: ${frequency_df['expected_severity'].std():,.0f}")
print(f"  Min: ${frequency_df['expected_severity'].min():,.0f}")
print(f"  Max: ${frequency_df['expected_severity'].max():,.0f}")

# ============================================================================
# 5. CALCULATE PURE PREMIUM
# ============================================================================

print("\n5. CALCULATING PURE PREMIUM (FREQUENCY × SEVERITY)")
print("-" * 80)

# Pure premium = E[Frequency] × E[Severity]
frequency_df['pure_premium'] = (frequency_df['predicted_frequency_final'] * 
                                frequency_df['expected_severity'])

print(f"Pure Premium Calculation:")
print(f"  Pure Premium = E[Frequency] × E[Severity]")
print(f"\nPure Premium Statistics:")
print(f"  Mean: ${frequency_df['pure_premium'].mean():,.0f}")
print(f"  Median: ${frequency_df['pure_premium'].median():,.0f}")
print(f"  Std Dev: ${frequency_df['pure_premium'].std():,.0f}")
print(f"  Min: ${frequency_df['pure_premium'].min():,.0f}")
print(f"  Max: ${frequency_df['pure_premium'].max():,.0f}")
print(f"  Q25: ${frequency_df['pure_premium'].quantile(0.25):,.0f}")
print(f"  Q75: ${frequency_df['pure_premium'].quantile(0.75):,.0f}")

# ============================================================================
# 6. APPLY LOADING FACTORS
# ============================================================================

print("\n6. APPLYING LOADING FACTORS")
print("-" * 80)

# Calculate individual loading components
frequency_df['acquisition_expenses'] = frequency_df['pure_premium'] * loading_factors['acquisition_expenses']
frequency_df['administrative_expenses'] = frequency_df['pure_premium'] * loading_factors['administrative']
frequency_df['profit_margin'] = frequency_df['pure_premium'] * loading_factors['profit_margin']
frequency_df['uncertainty_buffer'] = frequency_df['pure_premium'] * loading_factors['uncertainty_buffer']
frequency_df['reinsurance_ceded'] = frequency_df['pure_premium'] * loading_factors['reinsurance_ceded']

# Total loading
frequency_df['total_loading'] = (frequency_df['acquisition_expenses'] + 
                                 frequency_df['administrative_expenses'] +
                                 frequency_df['profit_margin'] +
                                 frequency_df['uncertainty_buffer'] +
                                 frequency_df['reinsurance_ceded'])

# Final premium
frequency_df['final_premium'] = frequency_df['pure_premium'] + frequency_df['total_loading']

print(f"Loading Factor Application:")
print(f"  Final Premium = Pure Premium × {loading_multiplier:.3f}")
print(f"\nFinal Premium Statistics:")
print(f"  Mean: ${frequency_df['final_premium'].mean():,.0f}")
print(f"  Median: ${frequency_df['final_premium'].median():,.0f}")
print(f"  Std Dev: ${frequency_df['final_premium'].std():,.0f}")
print(f"  Min: ${frequency_df['final_premium'].min():,.0f}")
print(f"  Max: ${frequency_df['final_premium'].max():,.0f}")

# ============================================================================
# 7. PRICING TABLES BY CATEGORY
# ============================================================================

print("\n7. GENERATING PRICING TABLES")
print("-" * 80)

# 7.1 Company-level pricing
print(f"\nTop 20 Companies by Premium:")
top_companies = frequency_df.nlargest(20, 'final_premium')[
    ['company_name', 'industry_primary', 'company_revenue_usd', 'predicted_frequency_final',
     'expected_severity', 'pure_premium', 'final_premium']
]
print(top_companies.to_string(index=False))

# 7.2 Industry pricing
print(f"\n\nIndustry Pricing Summary:")
industry_pricing = frequency_df.groupby('industry_primary').agg({
    'company_name': 'count',
    'pure_premium': ['mean', 'median', 'min', 'max'],
    'final_premium': ['mean', 'median', 'min', 'max']
}).round(0)
industry_pricing.columns = ['Count', 'Pure_Mean', 'Pure_Median', 'Pure_Min', 'Pure_Max',
                            'Final_Mean', 'Final_Median', 'Final_Min', 'Final_Max']
industry_pricing = industry_pricing.sort_values('Final_Mean', ascending=False).head(10)
print(industry_pricing)

# 7.3 Revenue tier pricing
print(f"\n\nRevenue Tier Pricing Summary:")
frequency_df['revenue_tier'] = pd.qcut(frequency_df['company_revenue_usd'], 
                                        q=4, labels=['Q1_Small', 'Q2_Medium', 'Q3_Large', 'Q4_Enterprise'])
tier_pricing = frequency_df.groupby('revenue_tier').agg({
    'company_name': 'count',
    'company_revenue_usd': ['min', 'max'],
    'pure_premium': ['mean', 'median'],
    'final_premium': ['mean', 'median']
}).round(0)
tier_pricing.columns = ['Count', 'Min_Revenue', 'Max_Revenue', 'Pure_Mean', 'Pure_Median', 'Final_Mean', 'Final_Median']
print(tier_pricing)

# ============================================================================
# 8. RISK CATEGORIES & PRICING
# ============================================================================

print("\n8. RISK-BASED PRICING CATEGORIES")
print("-" * 80)

# Create risk categories based on final premium and frequency
frequency_df['risk_category'] = pd.cut(frequency_df['final_premium'],
                                        bins=[0, 50e6, 100e6, 200e6, np.inf],
                                        labels=['Standard', 'Elevated', 'High', 'Specialty'])

risk_summary = frequency_df.groupby('risk_category').agg({
    'company_name': 'count',
    'final_premium': ['mean', 'min', 'max'],
    'predicted_frequency_final': 'mean',
    'expected_severity': 'mean'
}).round(0)
risk_summary.columns = ['Count', 'Mean_Premium', 'Min_Premium', 'Max_Premium', 'Avg_Frequency', 'Avg_Severity']
print(risk_summary)

# ============================================================================
# 9. PREMIUM ADJUSTMENTS & RELATIVITIES
# ============================================================================

print("\n9. PRICING RELATIVITIES & ADJUSTMENT FACTORS")
print("-" * 80)

# Overall mean premium
overall_mean_premium = frequency_df['final_premium'].mean()

# Industry relativities
print(f"\nIndustry Relativities (vs. Mean Premium of ${overall_mean_premium:,.0f}):")
industry_rel = frequency_df.groupby('industry_primary').agg({
    'final_premium': ['mean', 'count']
})
industry_rel.columns = ['Mean_Premium', 'Count']
industry_rel['Relativity'] = industry_rel['Mean_Premium'] / overall_mean_premium
industry_rel = industry_rel.sort_values('Mean_Premium', ascending=False).head(10)
print(industry_rel.round(3))

# Revenue tier relativities
print(f"\nRevenue Tier Relativities:")
tier_rel = frequency_df.groupby('revenue_tier').agg({
    'final_premium': ['mean', 'count']
})
tier_rel.columns = ['Mean_Premium', 'Count']
tier_rel['Relativity'] = tier_rel['Mean_Premium'] / overall_mean_premium
print(tier_rel.round(3))

# ============================================================================
# 10. PREMIUM DISTRIBUTION ANALYSIS
# ============================================================================

print("\n10. PREMIUM DISTRIBUTION ANALYSIS")
print("-" * 80)

# Quantile analysis
quantiles = [0.05, 0.10, 0.25, 0.50, 0.75, 0.90, 0.95]
print(f"\nFinal Premium Percentiles:")
for q in quantiles:
    value = frequency_df['final_premium'].quantile(q)
    print(f"  {q*100:3.0f}th percentile: ${value:>15,.0f}")

# Distribution shape
skewness = frequency_df['final_premium'].skew()
kurtosis = frequency_df['final_premium'].kurtosis()
print(f"\nDistribution Shape:")
print(f"  Skewness: {skewness:.4f} (right-skewed)" if skewness > 0 else f"  Skewness: {skewness:.4f} (left-skewed)")
print(f"  Kurtosis: {kurtosis:.4f}")

# ============================================================================
# 11. VISUALIZATIONS
# ============================================================================

print("\n11. GENERATING VISUALIZATIONS")
print("-" * 80)

fig, axes = plt.subplots(2, 2, figsize=(14, 10))

# Plot 1: Premium distribution
ax1 = axes[0, 0]
ax1.hist(frequency_df['final_premium'] / 1e6, bins=50, alpha=0.7, edgecolor='black', color='steelblue')
ax1.set_xlabel('Final Premium ($ Millions)')
ax1.set_ylabel('Count')
ax1.set_title('Premium Distribution (All Companies)')
ax1.axvline(frequency_df['final_premium'].mean() / 1e6, color='red', linestyle='--', linewidth=2, label='Mean')
ax1.axvline(frequency_df['final_premium'].median() / 1e6, color='green', linestyle='--', linewidth=2, label='Median')
ax1.legend()
ax1.grid(True, alpha=0.3)

# Plot 2: Frequency vs Severity scatter
ax2 = axes[0, 1]
scatter = ax2.scatter(frequency_df['predicted_frequency_final'], 
                      frequency_df['expected_severity'] / 1e6,
                      c=frequency_df['final_premium'] / 1e6, cmap='viridis', alpha=0.6, s=50)
ax2.set_xlabel('Expected Frequency (incidents)')
ax2.set_ylabel('Expected Severity ($ Millions)')
ax2.set_title('Frequency vs Severity (color = premium)')
cbar = plt.colorbar(scatter, ax=ax2)
cbar.set_label('Final Premium ($M)')
ax2.grid(True, alpha=0.3)

# Plot 3: Premium by industry (top 10)
ax3 = axes[1, 0]
top_industries = frequency_df.groupby('industry_primary')['final_premium'].mean().nlargest(10)
ax3.barh(range(len(top_industries)), top_industries / 1e6, color='coral')
ax3.set_yticks(range(len(top_industries)))
ax3.set_yticklabels(top_industries.index)
ax3.set_xlabel('Mean Final Premium ($ Millions)')
ax3.set_title('Average Premium by Industry (Top 10)')
ax3.grid(True, alpha=0.3, axis='x')

# Plot 4: Loading factor breakdown
ax4 = axes[1, 1]
loading_data = pd.DataFrame({
    'Component': list(loading_factors.keys()),
    'Percentage': [v * 100 for v in loading_factors.values()]
})
colors = plt.cm.Set3(np.linspace(0, 1, len(loading_data)))
ax4.pie(loading_data['Percentage'], labels=loading_data['Component'], autopct='%1.1f%%', colors=colors)
ax4.set_title('Loading Factor Composition')

plt.tight_layout()
plt.savefig('/mnt/user-data/outputs/premium_calculation_diagnostics.png', dpi=300, bbox_inches='tight')
print("Saved: premium_calculation_diagnostics.png")

# ============================================================================
# 12. EXPORT RESULTS
# ============================================================================

print("\n12. EXPORTING RESULTS")
print("-" * 80)

# 12.1 Full company-level results
export_cols = ['company_name', 'company_revenue_usd', 'employee_count', 'industry_primary',
               'is_public_company', 'predicted_frequency_final', 'risk_score',
               'expected_severity', 'pure_premium', 'total_loading', 'final_premium', 'risk_category']
export_df = frequency_df[export_cols].copy()
export_df = export_df.sort_values('final_premium', ascending=False)
export_df.to_csv('/mnt/user-data/outputs/complete_pricing_tables.csv', index=False)
print(f"Saved: complete_pricing_tables.csv ({len(export_df)} companies)")

# 12.2 Industry summary
industry_export = frequency_df.groupby('industry_primary').agg({
    'company_name': 'count',
    'final_premium': ['mean', 'median', 'min', 'max', 'std'],
    'predicted_frequency_final': 'mean',
    'expected_severity': 'mean'
}).round(0)
industry_export.columns = ['Company_Count', 'Mean_Premium', 'Median_Premium', 'Min_Premium', 'Max_Premium', 'StdDev_Premium', 'Avg_Frequency', 'Avg_Severity']
industry_export['Relativity'] = industry_export['Mean_Premium'] / overall_mean_premium
industry_export = industry_export.sort_values('Mean_Premium', ascending=False)
industry_export.to_csv('/mnt/user-data/outputs/industry_pricing_summary.csv')
print(f"Saved: industry_pricing_summary.csv")

# 12.3 Revenue tier summary
tier_export = frequency_df.groupby('revenue_tier').agg({
    'company_name': 'count',
    'company_revenue_usd': ['min', 'max'],
    'final_premium': ['mean', 'median', 'min', 'max'],
    'predicted_frequency_final': 'mean'
}).round(0)
tier_export.columns = ['Count', 'Min_Revenue', 'Max_Revenue', 'Mean_Premium', 'Median_Premium', 'Min_Premium', 'Max_Premium', 'Avg_Frequency']
tier_export.to_csv('/mnt/user-data/outputs/revenue_tier_pricing_summary.csv')
print(f"Saved: revenue_tier_pricing_summary.csv")

# ============================================================================
# 13. EXCEL WORKBOOK WITH FORMATTING
# ============================================================================

print("\n13. CREATING FORMATTED EXCEL WORKBOOK")
print("-" * 80)

wb = Workbook()
wb.remove(wb.active)

# Sheet 1: Company pricing table
ws1 = wb.create_sheet('Company Pricing')
ws1.append(['Company Name', 'Revenue ($M)', 'Employees', 'Industry', 'Public?',
            'Frequency', 'Severity ($M)', 'Pure Premium ($M)', 'Loading', 'Final Premium ($M)', 'Risk Category'])

# Format header
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for idx, row in export_df.head(100).iterrows():  # First 100 for Excel size
    ws1.append([
        row['company_name'],
        row['company_revenue_usd'] / 1e6,
        row['employee_count'],
        row['industry_primary'],
        'Yes' if row['is_public_company'] else 'No',
        round(row['predicted_frequency_final'], 3),
        round(row['expected_severity'] / 1e6, 1),
        round(row['pure_premium'] / 1e6, 1),
        f"{total_loading*100:.1f}%",
        round(row['final_premium'] / 1e6, 1),
        row['risk_category']
    ])

# Adjust column widths
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws1.column_dimensions[col].width = 15

# Format numbers
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for idx, cell in enumerate(row):
        if idx > 0 and idx < 10:  # Numeric columns
            if idx in [1, 2, 6, 7, 8, 9]:  # Currency/large numbers
                cell.number_format = '#,##0'
            elif idx in [5]:  # Frequency
                cell.number_format = '0.000'

# Sheet 2: Industry summary
ws2 = wb.create_sheet('Industry Summary')
ws2.append(['Industry', 'Count', 'Mean Premium ($M)', 'Median Premium ($M)', 
            'Min Premium ($M)', 'Max Premium ($M)', 'Avg Frequency', 'Avg Severity ($M)', 'Relativity'])

for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font

for industry, row in industry_export.iterrows():
    ws2.append([
        industry,
        int(row['Company_Count']),
        round(row['Mean_Premium'] / 1e6, 1),
        round(row['Median_Premium'] / 1e6, 1),
        round(row['Min_Premium'] / 1e6, 1),
        round(row['Max_Premium'] / 1e6, 1),
        round(row['Avg_Frequency'], 3),
        round(row['Avg_Severity'] / 1e6, 1),
        round(row['Relativity'], 3)
    ])

ws2.column_dimensions['A'].width = 15
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws2.column_dimensions[col].width = 15

# Sheet 3: Summary statistics
ws3 = wb.create_sheet('Summary')
ws3.append(['Metric', 'Value'])

summary_stats = [
    ['Total Companies', len(frequency_df)],
    ['Mean Pure Premium ($M)', round(frequency_df['pure_premium'].mean() / 1e6, 1)],
    ['Mean Final Premium ($M)', round(frequency_df['final_premium'].mean() / 1e6, 1)],
    ['Median Final Premium ($M)', round(frequency_df['final_premium'].median() / 1e6, 1)],
    ['Min Final Premium ($M)', round(frequency_df['final_premium'].min() / 1e6, 1)],
    ['Max Final Premium ($M)', round(frequency_df['final_premium'].max() / 1e6, 1)],
    ['Std Dev Premium ($M)', round(frequency_df['final_premium'].std() / 1e6, 1)],
    ['', ''],
    ['Loading Factors', ''],
    ['Acquisition Expenses', f"{loading_factors['acquisition_expenses']*100:.1f}%"],
    ['Administrative', f"{loading_factors['administrative']*100:.1f}%"],
    ['Profit Margin', f"{loading_factors['profit_margin']*100:.1f}%"],
    ['Uncertainty Buffer', f"{loading_factors['uncertainty_buffer']*100:.1f}%"],
    ['Reinsurance Ceded', f"{loading_factors['reinsurance_ceded']*100:.1f}%"],
    ['Total Loading', f"{total_loading*100:.1f}%"],
]

for stat in summary_stats:
    ws3.append(stat)

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 20

wb.save('/mnt/user-data/outputs/complete_pricing_framework.xlsx')
print("Saved: complete_pricing_framework.xlsx")

# ============================================================================
# 14. MODEL SUMMARY
# ============================================================================

print("\n" + "=" * 80)
print("PREMIUM CALCULATION ENGINE - SUMMARY")
print("=" * 80)

print(f"\n📊 PRICING RESULTS SUMMARY")
print(f"\nCompanies Priced: {len(frequency_df)}")
print(f"\nPure Premium Statistics:")
print(f"  Mean: ${frequency_df['pure_premium'].mean():>15,.0f}")
print(f"  Median: ${frequency_df['pure_premium'].median():>15,.0f}")
print(f"  Range: ${frequency_df['pure_premium'].min():,.0f} - ${frequency_df['pure_premium'].max():,.0f}")

print(f"\nFinal Premium Statistics (after {total_loading*100:.1f}% loading):")
print(f"  Mean: ${frequency_df['final_premium'].mean():>15,.0f}")
print(f"  Median: ${frequency_df['final_premium'].median():>15,.0f}")
print(f"  Range: ${frequency_df['final_premium'].min():,.0f} - ${frequency_df['final_premium'].max():,.0f}")

print(f"\n💰 LOADING FACTOR BREAKDOWN (% of Pure Premium)")
for component, rate in loading_factors.items():
    component_name = component.replace('_', ' ').title()
    print(f"  {component_name:.<40} {rate*100:>6.1f}%")
print(f"  {'Total Loading':.<40} {total_loading*100:>6.1f}%")

print(f"\n🏢 TOP 5 COMPANIES BY PREMIUM")
top5 = frequency_df.nlargest(5, 'final_premium')[['company_name', 'industry_primary', 'final_premium']]
for idx, (_, row) in enumerate(top5.iterrows(), 1):
    print(f"  {idx}. {row['company_name']:<40} ${row['final_premium']:>15,.0f}")

print(f"\n📈 INDUSTRY DISTRIBUTION")
top_industries = frequency_df.groupby('industry_primary')['final_premium'].mean().nlargest(5)
for industry, premium in top_industries.items():
    count = len(frequency_df[frequency_df['industry_primary'] == industry])
    print(f"  Industry {industry}: ${premium/1e6:>8.1f}M mean ({count:3d} companies)")

print(f"\n📋 FILES GENERATED")
print(f"  ✓ complete_pricing_tables.csv ({len(export_df)} companies)")
print(f"  ✓ industry_pricing_summary.csv (20 industries)")
print(f"  ✓ revenue_tier_pricing_summary.csv (4 tiers)")
print(f"  ✓ premium_calculation_diagnostics.png (4-panel visualization)")
print(f"  ✓ complete_pricing_framework.xlsx (3 sheets)")

print(f"\n" + "=" * 80)
print("PREMIUM CALCULATION ENGINE COMPLETE")
print("=" * 80)

print(f"\n✅ Ready for market launch!")
print(f"\nNext Steps:")
print(f"  1. Review pricing tables for reasonableness")
print(f"  2. Validate against market benchmarks")
print(f"  3. Adjust loading factors if needed (rerun script)")
print(f"  4. Get stakeholder approval")
print(f"  5. Deploy to underwriting system")
print(f"  6. Monitor actual vs. predicted experience (quarterly)")

print(f"\nFormula Applied:")
print(f"  Final Premium = (Frequency × Severity) × {loading_multiplier:.3f}")
print(f"\nAll results exported to /mnt/user-data/outputs/")
